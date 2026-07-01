#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Generate UX confirmation sheet (xlsx) from handoff documents.

Python 3.6+ required.
"""

import argparse
import io
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def emit(text):
    """Python 3 Unicode 安全输出。"""
    try:
        print(text)
    except UnicodeEncodeError:
        if hasattr(sys.stdout, "buffer"):
            sys.stdout.buffer.write(text.encode("utf-8", errors="replace") + b"\n")
        else:
            print(text.encode("ascii", errors="replace").decode("ascii"))


def read_file(path):
    try:
        with io.open(path, "r", encoding="utf-8") as fh:
            return fh.read()
    except Exception:
        return None


def _cell_value(value):
    """Pass-through cell value (Python 3 only)."""
    return value


def _relpath_or_abspath(path, base=None):
    """Return a relative path if under base, else the original path."""
    if base is None:
        base = os.getcwd()
    try:
        return os.path.relpath(path, start=base)
    except (ValueError, OSError):
        return path


# ---------------------------------------------------------------------------
# Markdown table parser for pending items
# ---------------------------------------------------------------------------

# Match the §15 待确认项 table header row
_PENDING_HEADER = re.compile(
    r"\|\s*编\s*号\s*\|.*\|\s*确\s*认\s*角\s*色\s*\|",
    re.UNICODE
)

# Row pattern: | 编号 | 问题 | 确认角色 | 跑偏风险 | 阻塞级别 | ... | 影响范围 |
_PENDING_ROW = re.compile(
    r"^\|\s*"
    r"(?P<id>\S+)\s*\|"
    r"(?P<question>[^|]+)\|"
    r"(?P<role>[^|]+)\|"
    r"(?P<deviation>[^|]+)\|"
    r"(?P<level>[^|]+)\|"
    r"(?P<options>[^|]+)\|"
    r"(?P<assumption>[^|]+)\|"
    r"(?P<timing>[^|]+)\|"
    r"(?P<impact>[^|]+)\|",
    re.UNICODE
)


def _extract_pending_items(handoff_path, include_non_blocking=False):
    """Extract pending items from a handoff markdown file.

    Args:
      handoff_path: path to the handoff markdown file.
      include_non_blocking: if True, include all levels (blocking + non-blocking).

    Returns:
      list of dict: each dict has keys matching the xlsx columns.
    """
    content = read_file(handoff_path)
    if not content:
        return []

    # Locate §15 待确认项 section
    section_start = -1
    lines = content.split("\n")
    for i, line in enumerate(lines):
        if _PENDING_HEADER.search(line):
            section_start = i + 2  # skip header + separator
            break

    if section_start == -1:
        return []

    items = []
    handoff_name = os.path.basename(handoff_path)

    # Extract H1 title from the handoff file
    handoff_title = handoff_name  # fallback
    for line in lines:
        if line.startswith("# "):
            handoff_title = line[2:].strip()
            break

    for line in lines[section_start:]:
        line = line.strip()
        if not line or not line.startswith("|"):
            # End of table
            if len(items) > 0 and not line.startswith("|"):
                break
            continue

        m = _PENDING_ROW.match(line)
        if not m:
            continue

        item_id = m.group("id").strip()
        question = m.group("question").strip()
        role = m.group("role").strip()
        # deviation = m.group("deviation").strip()
        level = m.group("level").strip()
        options = m.group("options").strip()
        assumption = m.group("assumption").strip()
        # timing = m.group("timing").strip()
        impact = m.group("impact").strip()

        # Skip resolved items (e.g. "~~blocking~~ → resolved")
        if "resolved" in level.lower():
            continue

        # Only include blocking / pre-blocking items (or all if --all)
        if not include_non_blocking and level not in ("blocking", "pre-blocking"):
            continue

        items.append({
            u"编号": item_id,
            u"来源 Handoff": handoff_name,
            u"文档标题": handoff_title,
            u"问题描述": question,
            u"阻塞级别": level,
            u"确认角色": ROLE_SHEET_MAP.get(role.strip().lower(), role),
            u"推荐选项": options,
            u"可继续假设": assumption,
            u"确认结论": u"",
            u"确认人": u"",
            u"确认日期": u"",
            u"_角色": role,
        })

    return items


def extract_all_pending(handoff_dir, include_non_blocking=False):
    """Traverse handoff_dir, extract pending items from all handoff files.

    Args:
      handoff_dir: directory containing handoff markdown files.
      include_non_blocking: if True, include non-blocking items.

    Returns:
      list of dict, each with pending item data.
    """
    all_items = []

    # Find the index file
    index_path = os.path.join(handoff_dir, "ux_handoff_index.md")
    if not os.path.isfile(index_path):
        emit("WARN: index file not found at {0}".format(index_path))

    for fname in os.listdir(handoff_dir):
        if not fname.endswith(".md"):
            continue
        if fname == "ux_handoff_index.md":
            continue
        fpath = os.path.join(handoff_dir, fname)
        items = _extract_pending_items(fpath, include_non_blocking=include_non_blocking)
        all_items.extend(items)

    return all_items


# ---------------------------------------------------------------------------
# Role grouping
# ---------------------------------------------------------------------------

# Map role values to Sheet names
ROLE_SHEET_MAP = {
    "product": u"产品经理",
    "interaction": u"交互设计师",
    "visual": u"视觉设计师",
    "technical": u"开发者",
    "host": u"宿主/平台",
    "data": u"数据/接口",
    "mixed": u"混合角色",
}


# ---------------------------------------------------------------------------
# xlsx generation
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
HEADER_FONT = Font(bold=True)
COLOR_MAP = {
    "pre-blocking": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "blocking": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
    "non-blocking": PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),
}
# Border styles
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_THICK_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)
# Highlight fill for mandatory columns (确认结论)
_FILLABLE_HEADER_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
_FILLABLE_CELL_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
# Bold font for mandatory rows
_BOLD_FONT = Font(bold=True)
_HEADER_BOLD_FONT = Font(bold=True, size=11)

COLUMNS = [u"编号", u"来源 Handoff", u"文档标题", u"问题描述", u"阻塞级别",
           u"确认角色", u"推荐选项", u"可继续假设", u"确认结论", u"确认人", u"确认日期"]
WIDE_COLS = {u"问题描述", u"推荐选项", u"可继续假设", u"确认结论"}
FILLABLE_COLS = {u"确认结论", u"确认人", u"确认日期"}  # columns to visually highlight


def _write_sheet_header(ws):
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=_cell_value(col_name))
        cell.font = _HEADER_BOLD_FONT
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
        # Highlight fillable columns
        if col_name in FILLABLE_COLS:
            cell.fill = _FILLABLE_HEADER_FILL
        else:
            cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def _write_sheet_rows(ws, items):
    for row_idx, item in enumerate(items, 2):
        level = item.get(u"阻塞级别", "").strip()
        is_mandatory = level in ("pre-blocking", "blocking")

        for col_idx, col_name in enumerate(COLUMNS, 1):
            value = item.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell_value(value))
            cell.alignment = Alignment(wrap_text=True, vertical="center")

            # Border for all data cells
            cell.border = _THIN_BORDER

            # Mandatory rows: bold text
            if is_mandatory:
                cell.font = _BOLD_FONT

            # Row-level color by blocking level (except fillable cells keep their own highlight)
            if col_name in FILLABLE_COLS:
                # Fillable columns: light green background, but keep blocking-level tint
                if is_mandatory:
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                else:
                    cell.fill = _FILLABLE_CELL_FILL
            else:
                fill = COLOR_MAP.get(level)
                if fill:
                    cell.fill = fill


def _set_column_widths(ws):
    for col_idx, col_name in enumerate(COLUMNS, 1):
        letter = get_column_letter(col_idx)
        if col_name in WIDE_COLS:
            ws.column_dimensions[letter].width = 45
        else:
            ws.column_dimensions[letter].width = 18


def _add_data_validation(ws, last_row):
    """Add dropdown validation for 阻塞级别 (E) and 确认角色 (F) columns."""
    # Blocking level dropdown on column E
    dv_level = DataValidation(
        type="list",
        formula1='"pre-blocking,blocking,non-blocking"',
        allow_blank=True
    )
    dv_level.error = u"请选择 pre-blocking / blocking / non-blocking"
    dv_level.errorTitle = u"无效值"
    ws.add_data_validation(dv_level)
    if last_row >= 2:
        dv_level.add("E2:E{0}".format(last_row))

    # Role dropdown on column F
    role_values = ",".join(sorted(ROLE_SHEET_MAP.values()))
    dv_role = DataValidation(
        type="list",
        formula1='"{0}"'.format(role_values),
        allow_blank=True
    )
    dv_role.error = u"请选择确认角色"
    dv_role.errorTitle = u"无效值"
    ws.add_data_validation(dv_role)
    if last_row >= 2:
        dv_role.add("F2:F{0}".format(last_row))


def _build_sheet0_meta(wb, output_dir, handoff_count, items):
    """Build Sheet 0「使用说明」with metadata and filling instructions."""
    ws = wb.create_sheet(title=u"使用说明", index=0)

    # Show relative paths for portability
    rel_dir = _relpath_or_abspath(output_dir) or "."
    index_rel = os.path.join(rel_dir, "ux_handoff_index.md")

    # Metadata section
    meta_rows = [
        [u"UX Confirmation Sheet 元数据", ""],
        [u"关联索引文件", _cell_value(index_rel)],
        [u"产出目录", _cell_value(rel_dir)],
        [u"关联 Handoff 数量", str(handoff_count)],
    ]

    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta_rows.append([u"生成时间", now])

    row = 1
    for label, value in meta_rows:
        ws.cell(row=row, column=1, value=_cell_value(label)).font = Font(bold=True)
        ws.cell(row=row, column=2, value=_cell_value(value))
        row += 1

    row += 1  # blank row

    # Overall summary
    blocking_count = sum(1 for it in items if it.get(u"阻塞级别", "").strip() in ("blocking", "pre-blocking"))
    ws.cell(row=row, column=1, value=_cell_value(u"待确认项总数")).font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(items))
    row += 1
    ws.cell(row=row, column=1, value=_cell_value(u"其中 blocking/pre-blocking")).font = Font(bold=True)
    ws.cell(row=row, column=2, value=blocking_count)
    row += 2  # blank row

    # Filling instructions
    ws.cell(row=row, column=1, value=_cell_value(u"填写约定")).font = Font(bold=True)
    row += 1
    instructions = [
        u"1. 所有待确认项在同一 Sheet「待确认项」中，各角色共同填写。可筛选「确认角色」列定位。",
        u"2. 阻塞级别与填写优先级：",
        u"   深红 pre-blocking → 必填，当前阶段阻塞，不确认无法启动",
        u"   浅红 blocking → 必填，开发启动前必须确认",
        u"   浅黄 non-blocking → 选填，可按「可继续假设」先行推进",
        u"3. 填写「确认结论」列：",
        u"   认同推荐选项 → 写「同推荐」或复制推荐选项内容",
        u"   有不同结论 → 自行描述最终决定",
        u"   ⚠️ 留空 → 默认按「可继续假设」执行",
        u"4. 同时填写「确认人」「确认日期」；如需改派，下拉切换「确认角色」即可。",
        u"5. 填写完成后放回原目录，对 AI 说「回放UX确认单」自动更新 handoff。",
    ]
    for instr in instructions:
        ws.cell(row=row, column=1, value=_cell_value(instr))
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    return ws


def generate_xlsx(items, output_dir, output_path):
    """Generate the confirmation sheet xlsx.

    Args:
      items: list of pending item dicts.
      output_dir: handoff output directory path.
      output_path: full path for the output xlsx.

    Returns:
      int: 0 on success, 1 on failure.
    """
    if openpyxl is None:
        emit("ERROR: openpyxl not installed. Install with: pip install openpyxl")
        return 1

    if not items:
        emit("INFO: No pending items found. No xlsx generated.")
        return 0

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Count unique handoff files
    handoff_files = set()
    for item in items:
        handoff_files.add(item.get(u"来源 Handoff", ""))

    # Build Sheet 0
    _build_sheet0_meta(wb, output_dir, len(handoff_files), items)

    # Build single data sheet
    ws = wb.create_sheet(title=u"待确认项")
    _write_sheet_header(ws)
    _write_sheet_rows(ws, items)
    _set_column_widths(ws)
    _add_data_validation(ws, len(items) + 1)

    try:
        wb.save(output_path)
        emit("INFO: Confirmation sheet saved to {0}".format(output_path))
        emit("INFO: {0} items in sheet「待确认项」.".format(len(items)))
    except Exception as e:
        emit("ERROR: Failed to save xlsx: {0}".format(str(e)))
        return 1

    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv):
    parser = argparse.ArgumentParser(
        description="Generate UX confirmation sheet (xlsx) from handoff documents."
    )
    parser.add_argument(
        "--handoff-dir", required=True,
        help="Directory containing handoff markdown files and index."
    )
    parser.add_argument(
        "--output",
        help="Output xlsx path (default: {handoff-dir}/ux_confirmation_sheet.xlsx)."
    )
    parser.add_argument(
        "--all", dest="include_all", action="store_true",
        help="Include non-blocking items as well (default: only blocking/pre-blocking)."
    )

    args = parser.parse_args(argv)

    handoff_dir = os.path.abspath(args.handoff_dir)
    if not os.path.isdir(handoff_dir):
        emit("ERROR: handoff directory not found: {0}".format(handoff_dir))
        return 1

    output_path = args.output
    if not output_path:
        output_path = os.path.join(handoff_dir, "ux_confirmation_sheet.xlsx")
    output_path = os.path.abspath(output_path)

    items = extract_all_pending(handoff_dir, include_non_blocking=args.include_all)

    if not items:
        emit("INFO: No pending items found in handoff files. Use --all to include non-blocking items.")
        emit("INFO: No xlsx generated.")
        return 0

    emit("INFO: Found {0} blocking/pre-blocking items across handoffs.".format(len(items)))
    return generate_xlsx(items, handoff_dir, output_path)


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))

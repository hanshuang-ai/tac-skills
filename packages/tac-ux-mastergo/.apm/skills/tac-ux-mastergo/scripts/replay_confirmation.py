#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Replay UX confirmation sheet(s) against handoff documents.

Reads one or more ux_confirmation_sheet*.xlsx files, matches confirmed items
to the originating handoff documents, detects merge conflicts when multiple
files contain different conclusions for the same item, and reports five
categories:

  adopted    — new confirmation, ready to write into handoff
  revised    — existing confirmation was changed, needs developer review
  unchanged  — already applied in handoff, no change
  conflict   — same item has different conclusions across xlsx files
  unresolved — conclusion column still empty

Python 3.6+ required.
"""

import argparse
import io
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def emit(text):
    """Python 3 Unicode 安全输出。"""
    try:
        print(text)
    except UnicodeEncodeError:
        if hasattr(sys.stdout, "buffer"):
            sys.stdout.buffer.write(text.encode("utf-8", errors="replace") + b"\n")
        else:
            print(text.encode("ascii", errors="replace").decode("ascii"))


def read_file(path):
    try:
        with io.open(path, "r", encoding="utf-8") as fh:
            return fh.read()
    except Exception:
        return None


def _safe_str(value):
    """Convert cell value to string (Python 3 only)."""
    if value is None:
        return ""
    return str(value)


# ---------------------------------------------------------------------------
# Handoff status extraction
# ---------------------------------------------------------------------------

# Match the §15 待确认项 table
_PENDING_HEADER_RE = re.compile(
    r"\|\s*编\s*号\s*\|.*\|\s*确\s*认\s*角\s*色\s*\|",
    re.UNICODE
)

_PENDING_ROW_RE = re.compile(
    r"^\|\s*"
    r"(?P<id>\S+)\s*\|"
    r"(?P<question>[^|]+)\|"
    r"(?P<role>[^|]+)\|"
    r"(?P<deviation>[^|]+)\|"
    r"(?P<level>[^|]+)\|"
    r"(?P<options>[^|]+)\|"
    r"(?P<assumption>[^|]+)\|"
    r"(?P<timing>[^|]+)\|"
    r"(?P<impact>[^|]+)\|",
    re.UNICODE
)

# Match an already-resolved row: ~~blocking~~ -> resolved or 已确认
_RESOLVED_RE = re.compile(r"~~.*~~\s*->\s*resolved|\u2705\s*\u5df2\u786e\u8ba4|resolved", re.UNICODE)


def read_handoff_pending_rows(handoff_dir):
    """Read all handoff files and extract current pending-item rows.

    Returns:
      dict: item_id -> {"结论": conclusion_text or "", "已确认": True/False, "handoff": filename}
    """
    current = {}

    for fname in os.listdir(handoff_dir):
        if not fname.endswith(".md"):
            continue
        if fname == "ux_handoff_index.md":
            continue

        fpath = os.path.join(handoff_dir, fname)
        content = read_file(fpath)
        if not content:
            continue

        lines = content.split("\n")
        section_start = -1
        for i, line in enumerate(lines):
            if _PENDING_HEADER_RE.search(line):
                section_start = i + 2
                break

        if section_start == -1:
            continue

        for line in lines[section_start:]:
            line_stripped = line.strip()
            if not line_stripped:
                continue
            if not line_stripped.startswith("|"):
                break

            m = _PENDING_ROW_RE.match(line_stripped)
            if not m:
                continue

            item_id = m.group("id").strip()
            # Extract the blocking level to detect resolution (scoped to level column)
            level_text = m.group("level").strip()
            # Try to detect if already resolved
            is_resolved = bool(_RESOLVED_RE.search(level_text))
            if "~~" in level_text:
                is_resolved = True

            # Try to extract existing conclusion from the row
            conclusion = ""
            if is_resolved:
                # Try to grab conclusion after the resolution marker
                after_resolve = line_stripped
                conclusion = _extract_conclusion_from_row(line_stripped)

            current[item_id] = {
                "已确认": is_resolved,
                "结论": conclusion,
                "handoff": fname,
            }

    return current


def _extract_conclusion_from_row(row_text):
    """Attempt to extract a resolution conclusion from an already-resolved row."""
    # Pattern: ✅ 已确认: text (确认人, 日期)
    m = re.search(r"[\u2705\U0001f7e2]\s*\u5df2\u786e\u8ba4[：:]\s*(.+?)(?:\s*\(|$)", row_text, re.UNICODE)
    if m:
        return m.group(1).strip()

    # Pattern: resolved: text
    m = re.search(r"resolved[：:]\s*(.+?)(?:\s*\(|$)", row_text, re.UNICODE | re.IGNORECASE)
    if m:
        return m.group(1).strip()

    # Pattern: -> resolved text
    m = re.search(r"->\s*resolved\s+(.+?)$", row_text, re.UNICODE | re.IGNORECASE)
    if m:
        return m.group(1).strip()

    return ""


# ---------------------------------------------------------------------------
# xlsx reading
# ---------------------------------------------------------------------------

def _find_xlsx_files(xlsx_dir):
    """Find all ux_confirmation_sheet*.xlsx files in the directory."""
    result = []
    for fname in os.listdir(xlsx_dir):
        if not fname.endswith(".xlsx"):
            continue
        if "ux_confirmation_sheet" not in fname:
            continue
        result.append(os.path.join(xlsx_dir, fname))
    return sorted(result)


def _read_xlsx_items(xlsx_path):
    """Read confirmation items from an xlsx file.

    Skips Sheet 0 (使用说明) and extracts data rows from role sheets.

    Returns:
      list of dict: each with keys 编号, 确认结论, 确认人, 确认日期, 阻塞级别,
                    _source (xlsx filename), _sheet (sheet name)
    """
    if openpyxl is None:
        emit("ERROR: openpyxl not installed. Install with: pip install openpyxl")
        return []

    items = []
    xlsx_name = os.path.basename(xlsx_path)

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        emit("WARN: Failed to read {0}: {1}".format(xlsx_path, str(e)))
        return []

    for sheet_name in wb.sheetnames:
        if sheet_name == u"使用说明":
            continue

        ws = wb[sheet_name]

        # Find header row
        header_map = {}
        data_start = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            row_values = [_safe_str(c) for c in row]
            if u"编号" in row_values:
                for ci, cv in enumerate(row_values):
                    header_map[cv] = ci
                data_start = row_idx + 1
                break

        if data_start is None:
            continue

        # Read data rows
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            row_values = [_safe_str(c).strip() for c in row]

            item_id = _cell(row_values, header_map.get(u"编号"))
            conclusion = _cell(row_values, header_map.get(u"确认结论"))
            if not item_id or not conclusion:
                continue

            person = _cell(row_values, header_map.get(u"确认人"))
            date = _cell(row_values, header_map.get(u"确认日期"))
            level = _cell(row_values, header_map.get(u"阻塞级别"))
            handoff_src = _cell(row_values, header_map.get(u"来源 Handoff"))
            impact = _cell(row_values, header_map.get(u"影响范围"))

            items.append({
                "编号": item_id,
                "确认结论": conclusion,
                "确认人": person,
                "确认日期": date,
                "阻塞级别": level,
                "_source": xlsx_name,
                "_sheet": sheet_name,
                "_handoff": handoff_src,
                "_影响范围": impact,
            })

    wb.close()
    return items


def _cell(values, index):
    if index is None or index >= len(values):
        return ""
    return values[index]


# ---------------------------------------------------------------------------
# Conflict detection
# ---------------------------------------------------------------------------

def classify_items(xlsx_items, handoff_status):
    """Classify each xlsx item against current handoff status.

    Returns five lists: adopted, revised, unchanged, conflict, unresolved.

    Conflict detection: same 编号 appears in multiple xlsx files with
    different 确认结论.
    """
    # Group xlsx items by 编号
    by_id = {}
    for item in xlsx_items:
        item_id = item["编号"]
        by_id.setdefault(item_id, []).append(item)

    adopted = []
    revised = []
    unchanged = []
    conflict = []
    unresolved = []

    for item_id, entries in sorted(by_id.items()):
        # Check for conflicts: same id, different conclusions across xlsx files
        conclusions = set(e["确认结论"] for e in entries)

        if len(entries) > 1 and len(conclusions) > 1:
            # Multi-file conflict
            conflict.append({
                "编号": item_id,
                "冲突详情": [
                    {"结论": e["确认结论"], "来源": e["_source"], "确认人": e["确认人"]}
                    for e in entries
                ],
                "_handoff": entries[0].get("_handoff", ""),
            })
            continue

        # Single conclusion (or multi-file same conclusion) — use the first
        entry = entries[0]
        conclusion = entry["确认结论"]

        # Check against current handoff status
        current = handoff_status.get(item_id)

        if current is None:
            # Item not found in handoff — treat as adopted (new)
            adopted.append(entry)
        elif not current["已确认"]:
            # Was not previously resolved — adopt
            adopted.append(entry)
        elif current["结论"] == conclusion:
            # Same conclusion, already applied — unchanged
            unchanged.append(entry)
        else:
            # Different conclusion — revised
            revised.append({
                "编号": item_id,
                "原结论": current["结论"],
                "新结论": conclusion,
                "确认人": entry["确认人"],
                "确认日期": entry["确认日期"],
                "_handoff": entry.get("_handoff", current.get("handoff", "")),
                "_source": entry["_source"],
            })

    # Also check handoff items that are still unresolved (no xlsx conclusion)
    for item_id, status in handoff_status.items():
        if not status["已确认"] and item_id not in by_id:
            unresolved.append({
                "编号": item_id,
                "阻塞级别": "unknown",
                "_handoff": status.get("handoff", ""),
            })

    return adopted, revised, unchanged, conflict, unresolved


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------

def _section(title, items, formatter, empty_msg="(none)"):
    emit(u"=== {0} ({1} items) ===".format(title, len(items)))
    if not items:
        emit(empty_msg)
    else:
        for item in items:
            emit(formatter(item))
    emit(u"")


def report(adopted, revised, unchanged, conflict, unresolved):
    """Print the five-category replay report."""

    # Summary
    total = len(adopted) + len(revised) + len(unchanged) + len(conflict) + len(unresolved)
    emit("=== REPLAY SUMMARY ===")
    emit("Total items: {0}".format(total))
    emit("  READY TO ADOPT:   {0}".format(len(adopted)))
    emit("  REVISED (review):  {0}".format(len(revised)))
    emit("  ALREADY APPLIED:   {0}".format(len(unchanged)))
    emit("  CONFLICT:          {0}".format(len(conflict)))
    emit("  UNRESOLVED:        {0}".format(len(unresolved)))
    emit("")

    # READY TO ADOPT
    _section(
        "READY TO ADOPT",
        adopted,
        lambda it: u"  {0} | {1} | {2} | 确认人: {3} | {4}".format(
            it.get(u"编号", u""),
            it.get(u"_handoff", u""),
            it.get(u"确认结论", u""),
            it.get(u"确认人", u""),
            it.get(u"确认日期", u""),
        )
    )

    # REVISED
    _section(
        u"REVISED - NEED REVIEW",
        revised,
        lambda it: u"  {0} | {1}\n    原结论: {2}\n    新结论: {3} | {4} | {5}".format(
            it.get(u"编号", u""),
            it.get(u"_handoff", u""),
            it.get(u"原结论", u""),
            it.get(u"新结论", u""),
            it.get(u"确认人", u""),
            it.get(u"确认日期", u""),
        )
    )

    # CONFLICT
    _section(
        u"CONFLICT - NEED RESOLUTION",
        conflict,
        lambda it: u"  {0} | {1}\n    {2}".format(
            it.get(u"编号", u""),
            it.get(u"_handoff", u""),
            u"\n    ".join(
                u"来自 {0}: {1} ({2})".format(
                    d.get(u"来源", u""), d.get(u"结论", u""), d.get(u"确认人", u"")
                )
                for d in it.get(u"冲突详情", [])
            ),
        )
    )

    # ALREADY APPLIED
    _section(
        u"ALREADY APPLIED - UNCHANGED",
        unchanged,
        lambda it: u"  {0} | {1} | {2}".format(
            it.get(u"编号", u""),
            it.get(u"_handoff", u""),
            it.get(u"确认结论", u""),
        )
    )

    # UNRESOLVED
    _section(
        u"UNRESOLVED",
        unresolved,
        lambda it: u"  {0} | {1}".format(
            it.get(u"编号", u""),
            it.get(u"_handoff", u""),
        )
    )

    # Action guidance
    emit("=== ACTION GUIDANCE ===")
    if adopted:
        emit("- {0} items ready to adopt: AI can write conclusions into handoff.".format(len(adopted)))
    if revised:
        emit("- {0} items revised: review changes before updating handoff.".format(len(revised)))
    if conflict:
        emit("- {0} conflicts: developer must resolve before handoff update.".format(len(conflict)))
    if unresolved:
        emit("- {0} items still unresolved: xlsx conclusion column is empty.".format(len(unresolved)))

    if not adopted and not revised and not conflict:
        if unresolved:
            emit("All xlsx items are still unresolved. Waiting for confirmation.")
        else:
            emit("All items already applied and unchanged. xlsx can be safely deleted.")

    # Return code: 0 if no blocking issues, 1 if conflicts/revised need attention
    if conflict or revised:
        return 1
    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv):
    parser = argparse.ArgumentParser(
        description="Replay UX confirmation sheet(s) against handoff documents."
    )
    parser.add_argument(
        "--xlsx-dir", required=True,
        help="Directory containing ux_confirmation_sheet*.xlsx files."
    )
    parser.add_argument(
        "--handoff-dir", required=True,
        help="Directory containing handoff markdown files and index."
    )
    parser.add_argument(
        "--json", action="store_true",
        help="Output results as JSON (not yet implemented)."
    )

    args = parser.parse_args(argv)

    if openpyxl is None:
        emit("ERROR: openpyxl not installed. Install with: pip install openpyxl")
        return 1

    xlsx_dir = os.path.abspath(args.xlsx_dir)
    handoff_dir = os.path.abspath(args.handoff_dir)

    if not os.path.isdir(xlsx_dir):
        emit("ERROR: xlsx directory not found: {0}".format(xlsx_dir))
        return 1

    if not os.path.isdir(handoff_dir):
        emit("ERROR: handoff directory not found: {0}".format(handoff_dir))
        return 1

    # Find xlsx files
    xlsx_files = _find_xlsx_files(xlsx_dir)
    if not xlsx_files:
        emit("ERROR: No ux_confirmation_sheet*.xlsx found in {0}".format(xlsx_dir))
        return 1

    emit("INFO: Found {0} xlsx file(s):".format(len(xlsx_files)))
    for f in xlsx_files:
        emit("  - {0}".format(os.path.basename(f)))

    # Read all xlsx items
    all_items = []
    for xf in xlsx_files:
        items = _read_xlsx_items(xf)
        all_items.extend(items)
        emit("INFO: {0} -> {1} items with conclusions".format(
            os.path.basename(xf), len(items)))

    if not all_items:
        emit("INFO: No filled conclusions found in any xlsx. Nothing to replay.")
        return 0

    # Read current handoff status
    handoff_status = read_handoff_pending_rows(handoff_dir)
    emit("INFO: {0} pending items found in handoffs.".format(len(handoff_status)))

    # Classify
    adopted, revised, unchanged, conflict, unresolved = classify_items(
        all_items, handoff_status
    )

    # Report
    return report(adopted, revised, unchanged, conflict, unresolved)


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
